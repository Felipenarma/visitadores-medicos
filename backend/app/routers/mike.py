from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from datetime import datetime, timedelta
from typing import Any, Optional, List
import os
import json
import uuid
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import anthropic
import openpyxl
from ..database import get_db
from ..models import Visit, Doctor, MedicalRep, Sale, BusinessLine, MikeMemory
from ..schemas import AgentMessage

router = APIRouter(prefix="/api/mike", tags=["mike"])

# Module-level store for Excel exports keyed by token
_export_store: dict = {}

MIKE_SYSTEM_PROMPT = """Eres Mike, el asistente de IA ejecutivo del laboratorio Narma para el administrador Felipe.

Tu rol principal es:
- Analizar y gestionar el desempeño de los visitadores médicos (visitas, tasa de cumplimiento, comportamiento, tendencias)
- Gestionar calendarios y visitas: programar, reagendar y cancelar visitas de cualquier visitador
- Asignar médicos a visitadores: por línea de productos, especialidad, zona o criterio que decidas
- Analizar la cartera de médicos (nuevos, ranking, médicos sin visitar, inactivos)
- Analizar la cartera de pacientes (ranking por volumen, nuevos pacientes, retención, historial por paciente, qué médicos los atienden)
- Obtener métricas de ventas por período, línea de negocio, visitador, médico o paciente
- Calcular y revisar comisiones
- Identificar oportunidades y riesgos: médicos que dejaron de comprar, pacientes que no han comprado recientemente, visitadores con bajo rendimiento, médicos sin asignar

Cuando analices el comportamiento de un visitador, examina:
- Tasa de cumplimiento de visitas (completadas vs programadas)
- Días de la semana que más trabaja y a qué horas
- Tendencia mensual: ¿está mejorando o empeorando?
- Comparación con otros visitadores del equipo
- Médicos que visita vs médicos asignados (cobertura real)
- Relación entre visitas y ventas generadas

Responde siempre en español. Sé analítico, preciso y proactivo. Si detectas algo relevante en los datos, coméntalo sin que te lo pidan.
Cuando el usuario pida exportar datos a Excel, usa export_to_excel.

Fecha actual: {current_date}"""

MIKE_TOOLS = [
    {
        "name": "get_dashboard_overview",
        "description": "Obtiene KPIs generales de la plataforma: total médicos activos, visitadores activos, visitas de hoy/semana/mes y distribución de ventas por línea de negocio.",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "get_all_reps_summary",
        "description": "Resumen de todos los visitadores: nombre, médicos asignados, visitas del mes, tasa de cumplimiento y ventas generadas.",
        "input_schema": {
            "type": "object",
            "properties": {
                "month": {"type": "integer", "description": "Mes (1-12). Si no se indica, usa el mes actual."},
                "year": {"type": "integer", "description": "Año (ej: 2026). Si no se indica, usa el año actual."}
            }
        }
    },
    {
        "name": "get_rep_detail",
        "description": "Análisis detallado de un visitador específico: visitas realizadas vs programadas, médicos asignados, ventas por categoría, médicos nuevos y comisión estimada.",
        "input_schema": {
            "type": "object",
            "properties": {
                "rep_id": {"type": "integer", "description": "ID del visitador"},
                "month": {"type": "integer", "description": "Mes (1-12)"},
                "year": {"type": "integer", "description": "Año"}
            },
            "required": ["rep_id"]
        }
    },
    {
        "name": "get_doctor_ranking",
        "description": "Ranking de médicos por volumen de ventas (unidades o monto). Útil para identificar los mejores prescriptores.",
        "input_schema": {
            "type": "object",
            "properties": {
                "month": {"type": "integer", "description": "Mes (1-12)"},
                "year": {"type": "integer", "description": "Año"},
                "top": {"type": "integer", "description": "Cuántos médicos mostrar (default: 20)"},
                "rep_id": {"type": "integer", "description": "Filtrar por visitador específico (opcional)"},
                "business_line_id": {"type": "integer", "description": "Filtrar por línea de negocio (opcional)"}
            }
        }
    },
    {
        "name": "get_new_doctors",
        "description": "Médicos que compraron por primera vez en un período determinado. Indicador clave de crecimiento de cartera.",
        "input_schema": {
            "type": "object",
            "properties": {
                "month": {"type": "integer", "description": "Mes (1-12)"},
                "year": {"type": "integer", "description": "Año"}
            }
        }
    },
    {
        "name": "get_sales_analysis",
        "description": "Análisis de ventas con múltiples filtros. Puede agrupar por médico, visitador o línea de negocio. Compara períodos.",
        "input_schema": {
            "type": "object",
            "properties": {
                "month": {"type": "integer", "description": "Mes (1-12)"},
                "year": {"type": "integer", "description": "Año"},
                "rep_id": {"type": "integer", "description": "Filtrar por visitador (opcional)"},
                "doctor_id": {"type": "integer", "description": "Filtrar por médico (opcional)"},
                "business_line_name": {"type": "string", "description": "Filtrar por nombre de línea de negocio (opcional)"},
                "group_by": {
                    "type": "string",
                    "enum": ["doctor", "rep", "business_line", "month", "category"],
                    "description": "Cómo agrupar los resultados"
                }
            }
        }
    },
    {
        "name": "get_rep_commissions",
        "description": "Comisiones calculadas por visitador para un período: ventas totales, médicos nuevos, detalle por médico y categoría.",
        "input_schema": {
            "type": "object",
            "properties": {
                "month": {"type": "integer", "description": "Mes (1-12)"},
                "year": {"type": "integer", "description": "Año"},
                "rep_id": {"type": "integer", "description": "Filtrar por visitador específico (opcional)"}
            }
        }
    },
    {
        "name": "get_visits_tracking",
        "description": "Seguimiento de visitas: completitud por visitador para una fecha o rango. Identifica visitadores que no están cumpliendo sus visitas.",
        "input_schema": {
            "type": "object",
            "properties": {
                "date": {"type": "string", "description": "Fecha en formato YYYY-MM-DD (default: hoy)"},
                "month": {"type": "integer", "description": "Si se indica, analiza todo el mes en lugar de un día"},
                "year": {"type": "integer", "description": "Año para análisis mensual"}
            }
        }
    },
    {
        "name": "search_doctors",
        "description": "Busca médicos por nombre, especialidad o visitador asignado. Devuelve datos de contacto, asignación y última visita.",
        "input_schema": {
            "type": "object",
            "properties": {
                "search": {"type": "string", "description": "Nombre parcial del médico"},
                "specialty": {"type": "string", "description": "Especialidad médica"},
                "rep_id": {"type": "integer", "description": "Filtrar por visitador"},
                "without_rep": {"type": "boolean", "description": "Si es true, solo médicos sin visitador asignado"},
                "limit": {"type": "integer", "description": "Máximo de resultados (default: 30)"}
            }
        }
    },
    {
        "name": "get_doctor_detail",
        "description": "Información completa de un médico: datos de contacto, visitador asignado, historial de visitas, ventas históricas y evolución mensual.",
        "input_schema": {
            "type": "object",
            "properties": {
                "doctor_id": {"type": "integer", "description": "ID del médico"}
            },
            "required": ["doctor_id"]
        }
    },
    {
        "name": "get_inactive_doctors",
        "description": "Médicos que no han recibido visitas o no han comprado en los últimos N días. Útil para detectar cartera en riesgo.",
        "input_schema": {
            "type": "object",
            "properties": {
                "days_without_visit": {"type": "integer", "description": "Días sin visita para considerar inactivo (default: 45)"},
                "rep_id": {"type": "integer", "description": "Filtrar por visitador (opcional)"}
            }
        }
    },
    {
        "name": "list_reps",
        "description": "Lista todos los visitadores médicos con sus datos básicos (id, nombre, email, teléfono, zona).",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "assign_rep_to_doctor",
        "description": "Asigna o cambia el visitador asignado a un médico.",
        "input_schema": {
            "type": "object",
            "properties": {
                "doctor_id": {"type": "integer", "description": "ID del médico"},
                "rep_id": {"type": "integer", "description": "ID del visitador. Usar 0 para desasignar."}
            },
            "required": ["doctor_id", "rep_id"]
        }
    },
    {
        "name": "get_monthly_trend",
        "description": "Evolución mensual de ventas y visitas para los últimos N meses. Ideal para ver tendencias de crecimiento.",
        "input_schema": {
            "type": "object",
            "properties": {
                "months": {"type": "integer", "description": "Cuántos meses hacia atrás analizar (default: 6)"},
                "rep_id": {"type": "integer", "description": "Filtrar por visitador (opcional)"},
                "business_line_name": {"type": "string", "description": "Filtrar por línea de negocio (opcional)"}
            }
        }
    },
    {
        "name": "get_rep_calendar",
        "description": "Obtiene el calendario de visitas de un visitador para un rango de fechas. Muestra visitas programadas, completadas y perdidas organizadas por día.",
        "input_schema": {
            "type": "object",
            "properties": {
                "rep_id": {"type": "integer", "description": "ID del visitador"},
                "date_from": {"type": "string", "description": "Fecha inicio en formato YYYY-MM-DD (default: hoy)"},
                "date_to": {"type": "string", "description": "Fecha fin en formato YYYY-MM-DD (default: 30 días adelante)"},
                "status": {"type": "string", "description": "Filtrar por estado: scheduled, completed, missed, cancelled (opcional)"}
            },
            "required": ["rep_id"]
        }
    },
    {
        "name": "schedule_visit",
        "description": "Programa una nueva visita para un visitador con un médico específico.",
        "input_schema": {
            "type": "object",
            "properties": {
                "rep_id": {"type": "integer", "description": "ID del visitador"},
                "doctor_id": {"type": "integer", "description": "ID del médico a visitar"},
                "scheduled_date": {"type": "string", "description": "Fecha de la visita en formato YYYY-MM-DD"},
                "notes": {"type": "string", "description": "Notas opcionales para la visita"}
            },
            "required": ["rep_id", "doctor_id", "scheduled_date"]
        }
    },
    {
        "name": "reschedule_visit",
        "description": "Reagenda una visita existente a otra fecha.",
        "input_schema": {
            "type": "object",
            "properties": {
                "visit_id": {"type": "integer", "description": "ID de la visita"},
                "new_date": {"type": "string", "description": "Nueva fecha en formato YYYY-MM-DD"},
                "notes": {"type": "string", "description": "Notas opcionales"}
            },
            "required": ["visit_id", "new_date"]
        }
    },
    {
        "name": "cancel_visit",
        "description": "Cancela una visita programada.",
        "input_schema": {
            "type": "object",
            "properties": {
                "visit_id": {"type": "integer", "description": "ID de la visita a cancelar"},
                "reason": {"type": "string", "description": "Motivo de cancelación (opcional)"}
            },
            "required": ["visit_id"]
        }
    },
    {
        "name": "bulk_assign_doctors",
        "description": "Asigna múltiples médicos a un visitador según criterios: por línea de negocio, especialidad, ciudad, o médicos sin asignar. Útil para organizar la cartera de un visitador.",
        "input_schema": {
            "type": "object",
            "properties": {
                "rep_id": {"type": "integer", "description": "ID del visitador al que asignar los médicos"},
                "business_line_name": {"type": "string", "description": "Asignar médicos de esta línea de negocio (ej: 'Hormonas', 'Dermatología')"},
                "specialty": {"type": "string", "description": "Asignar médicos de esta especialidad"},
                "city": {"type": "string", "description": "Asignar médicos de esta ciudad"},
                "without_rep_only": {"type": "boolean", "description": "Si es true, solo asigna médicos que no tienen visitador (default: true)"},
                "limit": {"type": "integer", "description": "Máximo de médicos a asignar (default: 50)"},
                "doctor_ids": {"type": "array", "items": {"type": "integer"}, "description": "Lista específica de IDs de médicos a asignar (opcional, tiene prioridad sobre filtros)"}
            },
            "required": ["rep_id"]
        }
    },
    {
        "name": "get_rep_behavior_analysis",
        "description": "Análisis profundo del comportamiento de un visitador: patrones de trabajo, días activos, horarios, tasa de cumplimiento por semana, comparación con el equipo, correlación visitas-ventas.",
        "input_schema": {
            "type": "object",
            "properties": {
                "rep_id": {"type": "integer", "description": "ID del visitador a analizar"},
                "months": {"type": "integer", "description": "Cuántos meses de historial analizar (default: 3)"}
            },
            "required": ["rep_id"]
        }
    },
    {
        "name": "save_to_memory",
        "description": "Guarda un hecho, decisión o conclusión importante en la memoria persistente de Mike. Úsala al final de conversaciones relevantes para recordar acuerdos, patrones detectados, cambios de estrategia o cualquier dato importante para futuras sesiones.",
        "input_schema": {
            "type": "object",
            "properties": {
                "content": {"type": "string", "description": "El hecho o conclusión a recordar (ej: 'Angelo tiene 436 médicos activos, 43 reasignados a Marco en julio 2026')"},
                "category": {
                    "type": "string",
                    "enum": ["general", "visitador", "medico", "venta", "decision", "alerta"],
                    "description": "Categoría del recuerdo"
                }
            },
            "required": ["content"]
        }
    },
    {
        "name": "get_memories",
        "description": "Recupera todo lo guardado en la memoria persistente de Mike. Úsala cuando necesites recordar conversaciones anteriores, decisiones pasadas o contexto histórico.",
        "input_schema": {
            "type": "object",
            "properties": {
                "category": {"type": "string", "description": "Filtrar por categoría (opcional)"}
            }
        }
    },
    {
        "name": "delete_memory",
        "description": "Elimina un recuerdo obsoleto o incorrecto de la memoria de Mike.",
        "input_schema": {
            "type": "object",
            "properties": {
                "memory_id": {"type": "integer", "description": "ID del recuerdo a eliminar"}
            },
            "required": ["memory_id"]
        }
    },
    {
        "name": "assign_business_line_to_doctor",
        "description": "Asigna una línea de negocio a un médico específico. Usa esto cuando el usuario quiera cambiar o establecer la línea de un médico en particular.",
        "input_schema": {
            "type": "object",
            "properties": {
                "doctor_id": {"type": "integer", "description": "ID del médico"},
                "business_line_name": {"type": "string", "description": "Nombre de la línea de negocio (ej: 'Hormonas', 'Dermatología', 'Cannabis Medicinal', 'Control de Peso', 'Suero Terapia', 'Veterinaria')"}
            },
            "required": ["doctor_id", "business_line_name"]
        }
    },
    {
        "name": "auto_assign_business_lines",
        "description": "Asigna automáticamente líneas de negocio a médicos basándose en las categorías de sus prescripciones/ventas. Analiza las ventas de cada médico, determina la categoría dominante y la mapea a la línea de negocio correspondiente. Puede aplicarse a todos los médicos sin línea asignada, o a todos.",
        "input_schema": {
            "type": "object",
            "properties": {
                "only_without_line": {"type": "boolean", "description": "Si es true (default), solo asigna a médicos que aún no tienen línea. Si es false, reasigna a todos según sus ventas actuales."},
                "rep_id": {"type": "integer", "description": "Limitar a médicos de un visitador específico (opcional)"},
                "dry_run": {"type": "boolean", "description": "Si es true, solo muestra qué se asignaría sin hacer cambios (preview). Default: false."}
            }
        }
    },
    {
        "name": "get_patient_analysis",
        "description": "Analiza la cartera de pacientes: ranking de los que más compran, pacientes nuevos en un período, pacientes que no han comprado en N días (retención), y distribución por producto/categoría. Úsalo cuando el usuario pregunte por pacientes, quiera ver quiénes son los mejores clientes, o detectar pacientes en riesgo de abandono.",
        "input_schema": {
            "type": "object",
            "properties": {
                "month": {"type": "integer", "description": "Mes (1-12). Si no se indica, usa el mes actual."},
                "year": {"type": "integer", "description": "Año. Si no se indica, usa el año actual."},
                "top": {"type": "integer", "description": "Cuántos pacientes mostrar en el ranking (default: 20)"},
                "analysis_type": {
                    "type": "string",
                    "enum": ["ranking", "new_patients", "retention", "by_category"],
                    "description": "Tipo de análisis: ranking (top pacientes por volumen), new_patients (pacientes nuevos en el período), retention (pacientes que no han comprado en N días), by_category (distribución por categoría de producto)"
                },
                "days_without_purchase": {"type": "integer", "description": "Para retention: días sin compra para considerar en riesgo (default: 60)"},
                "doctor_id": {"type": "integer", "description": "Filtrar pacientes de un médico específico (opcional)"},
                "rep_id": {"type": "integer", "description": "Filtrar por visitador (opcional)"},
                "categoria": {"type": "string", "description": "Filtrar por categoría de producto (opcional)"}
            }
        }
    },
    {
        "name": "get_patient_detail",
        "description": "Información completa de un paciente específico: historial de compras, productos que recibe, médicos que lo atienden, evolución mensual de gasto. Úsalo cuando el usuario quiera saber todo sobre un paciente en particular.",
        "input_schema": {
            "type": "object",
            "properties": {
                "rut_paciente": {"type": "string", "description": "RUT del paciente (con o sin guión/puntos)"},
                "nombre_paciente": {"type": "string", "description": "Nombre parcial del paciente (alternativa al RUT)"}
            }
        }
    },
    {
        "name": "export_to_excel",
        "description": "Exporta datos a un archivo Excel descargable. Úsalo cuando el usuario pida exportar, descargar o generar un Excel.",
        "input_schema": {
            "type": "object",
            "properties": {
                "data_type": {
                    "type": "string",
                    "enum": ["ranking_medicos", "visitadores", "nuevos_medicos", "tendencia_ventas", "comisiones", "ranking_pacientes", "pacientes_nuevos"],
                    "description": "Tipo de datos a exportar"
                },
                "month": {"type": "integer", "description": "Mes (1-12). Si no se indica, usa el mes actual."},
                "year": {"type": "integer", "description": "Año. Si no se indica, usa el año actual."},
                "rep_id": {"type": "integer", "description": "Filtrar por visitador específico (opcional)"}
            },
            "required": ["data_type"]
        }
    }
]


def execute_mike_tool(tool_name: str, tool_input: dict, db: Session) -> Any:
    now = datetime.utcnow()
    current_month = now.month
    current_year = now.year

    # ── get_dashboard_overview ────────────────────────────────────────────────
    if tool_name == "get_dashboard_overview":
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today_start + timedelta(days=1)
        week_start = today_start - timedelta(days=today_start.weekday())
        month_start = today_start.replace(day=1)

        total_doctors = db.query(func.count(Doctor.id)).filter(Doctor.is_active == True).scalar()
        total_reps = db.query(func.count(MedicalRep.id)).filter(MedicalRep.is_active == True).scalar()

        visits_today = db.query(func.count(Visit.id)).filter(
            Visit.scheduled_date >= today_start, Visit.scheduled_date < today_end
        ).scalar()
        visits_completed_today = db.query(func.count(Visit.id)).filter(
            Visit.scheduled_date >= today_start, Visit.scheduled_date < today_end,
            Visit.status == "completed"
        ).scalar()
        visits_week = db.query(func.count(Visit.id)).filter(
            Visit.scheduled_date >= week_start, Visit.scheduled_date < today_end
        ).scalar()
        visits_month = db.query(func.count(Visit.id)).filter(
            Visit.scheduled_date >= month_start
        ).scalar()
        visits_completed_month = db.query(func.count(Visit.id)).filter(
            Visit.scheduled_date >= month_start, Visit.status == "completed"
        ).scalar()

        # Sales this month
        sales_month = db.query(func.sum(Sale.amount)).filter(
            extract('month', Sale.sale_date) == current_month,
            extract('year', Sale.sale_date) == current_year
        ).scalar() or 0

        sales_count_month = db.query(func.count(Sale.id)).filter(
            extract('month', Sale.sale_date) == current_month,
            extract('year', Sale.sale_date) == current_year
        ).scalar()

        # By business line this month
        bl_sales = db.query(
            BusinessLine.name, func.sum(Sale.amount).label("total"), func.count(Sale.id).label("count")
        ).join(Doctor, Doctor.id == Sale.doctor_id, isouter=True
        ).join(BusinessLine, BusinessLine.id == Doctor.business_line_id, isouter=True
        ).filter(
            extract('month', Sale.sale_date) == current_month,
            extract('year', Sale.sale_date) == current_year
        ).group_by(BusinessLine.name).all()

        return {
            "fecha": now.strftime("%d/%m/%Y"),
            "medicos_activos": total_doctors,
            "visitadores_activos": total_reps,
            "visitas_hoy": {"programadas": visits_today, "completadas": visits_completed_today},
            "visitas_semana": visits_week,
            "visitas_mes": {"programadas": visits_month, "completadas": visits_completed_month,
                            "tasa_cumplimiento": round(visits_completed_month / visits_month * 100, 1) if visits_month else 0},
            "ventas_mes": {"total_monto": round(float(sales_month), 2), "total_registros": sales_count_month},
            "ventas_por_linea": [{"linea": r.name or "Sin línea", "monto": round(float(r.total or 0), 2), "registros": r.count} for r in bl_sales]
        }

    # ── get_all_reps_summary ──────────────────────────────────────────────────
    elif tool_name == "get_all_reps_summary":
        month = tool_input.get("month", current_month)
        year = tool_input.get("year", current_year)
        month_start = datetime(year, month, 1)
        if month == 12:
            month_end = datetime(year + 1, 1, 1)
        else:
            month_end = datetime(year, month + 1, 1)

        reps = db.query(MedicalRep).filter(MedicalRep.is_active == True).all()
        result = []
        for rep in reps:
            doctors_count = db.query(func.count(Doctor.id)).filter(
                Doctor.rep_id == rep.id, Doctor.is_active == True
            ).scalar()

            visits_programmed = db.query(func.count(Visit.id)).filter(
                Visit.rep_id == rep.id,
                Visit.scheduled_date >= month_start,
                Visit.scheduled_date < month_end
            ).scalar()
            visits_done = db.query(func.count(Visit.id)).filter(
                Visit.rep_id == rep.id,
                Visit.scheduled_date >= month_start,
                Visit.scheduled_date < month_end,
                Visit.status == "completed"
            ).scalar()

            sales_amount = db.query(func.sum(Sale.amount)).join(
                Doctor, Doctor.id == Sale.doctor_id
            ).filter(
                Doctor.rep_id == rep.id,
                extract('month', Sale.sale_date) == month,
                extract('year', Sale.sale_date) == year
            ).scalar() or 0

            sales_count = db.query(func.count(Sale.id)).join(
                Doctor, Doctor.id == Sale.doctor_id
            ).filter(
                Doctor.rep_id == rep.id,
                extract('month', Sale.sale_date) == month,
                extract('year', Sale.sale_date) == year
            ).scalar()

            result.append({
                "rep_id": rep.id,
                "nombre": rep.name,
                "zona": rep.zone,
                "territorio": rep.territory,
                "medicos_asignados": doctors_count,
                "visitas_programadas": visits_programmed,
                "visitas_completadas": visits_done,
                "tasa_cumplimiento": round(visits_done / visits_programmed * 100, 1) if visits_programmed else 0,
                "ventas_monto": round(float(sales_amount), 2),
                "ventas_registros": sales_count
            })

        result.sort(key=lambda x: x["ventas_monto"], reverse=True)
        return {"periodo": f"{month:02d}/{year}", "visitadores": result, "total": len(result)}

    # ── get_rep_detail ────────────────────────────────────────────────────────
    elif tool_name == "get_rep_detail":
        rep_id = tool_input.get("rep_id")
        month = tool_input.get("month", current_month)
        year = tool_input.get("year", current_year)

        rep = db.query(MedicalRep).filter(MedicalRep.id == rep_id).first()
        if not rep:
            return {"error": f"Visitador {rep_id} no encontrado"}

        month_start = datetime(year, month, 1)
        month_end = datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1)

        # Visits
        visits = db.query(Visit).filter(
            Visit.rep_id == rep_id,
            Visit.scheduled_date >= month_start,
            Visit.scheduled_date < month_end
        ).all()
        completed = [v for v in visits if v.status == "completed"]
        missed = [v for v in visits if v.status == "missed"]

        # Doctors assigned
        doctors = db.query(Doctor).filter(Doctor.rep_id == rep_id, Doctor.is_active == True).all()

        # Sales by category
        sales_by_cat = db.query(
            Sale.categoria, func.sum(Sale.amount).label("total"), func.count(Sale.id).label("count")
        ).join(Doctor, Doctor.id == Sale.doctor_id
        ).filter(
            Doctor.rep_id == rep_id,
            extract('month', Sale.sale_date) == month,
            extract('year', Sale.sale_date) == year
        ).group_by(Sale.categoria).all()

        total_sales = sum(float(r.total or 0) for r in sales_by_cat)

        # New doctors (first sale ever)
        new_docs = []
        for doc in doctors:
            first_sale = db.query(Sale).filter(Sale.doctor_id == doc.id).order_by(Sale.sale_date.asc()).first()
            if first_sale and first_sale.sale_date:
                if first_sale.sale_date.month == month and first_sale.sale_date.year == year:
                    new_docs.append({"doctor_id": doc.id, "nombre": doc.name, "especialidad": doc.specialty})

        return {
            "rep_id": rep.id,
            "nombre": rep.name,
            "zona": rep.zone,
            "periodo": f"{month:02d}/{year}",
            "medicos_asignados": len(doctors),
            "visitas": {
                "programadas": len(visits),
                "completadas": len(completed),
                "perdidas": len(missed),
                "tasa_cumplimiento": round(len(completed) / len(visits) * 100, 1) if visits else 0
            },
            "ventas_total": round(total_sales, 2),
            "ventas_por_categoria": [
                {"categoria": r.categoria or "Sin categoría", "monto": round(float(r.total or 0), 2), "registros": r.count}
                for r in sorted(sales_by_cat, key=lambda x: float(x.total or 0), reverse=True)
            ],
            "medicos_nuevos": new_docs,
            "lista_medicos": [{"id": d.id, "nombre": d.name, "especialidad": d.specialty} for d in doctors[:20]]
        }

    # ── get_doctor_ranking ────────────────────────────────────────────────────
    elif tool_name == "get_doctor_ranking":
        month = tool_input.get("month", current_month)
        year = tool_input.get("year", current_year)
        top = tool_input.get("top", 20)
        rep_id = tool_input.get("rep_id")
        business_line_id = tool_input.get("business_line_id")

        query = db.query(
            Doctor.id, Doctor.name, Doctor.specialty,
            func.sum(Sale.amount).label("total_monto"),
            func.count(Sale.id).label("total_registros")
        ).join(Sale, Sale.doctor_id == Doctor.id
        ).filter(
            extract('month', Sale.sale_date) == month,
            extract('year', Sale.sale_date) == year
        )
        if rep_id:
            query = query.filter(Doctor.rep_id == rep_id)
        if business_line_id:
            query = query.filter(Doctor.business_line_id == business_line_id)

        rows = query.group_by(Doctor.id, Doctor.name, Doctor.specialty
                              ).order_by(func.sum(Sale.amount).desc()).limit(top).all()

        # Previous month for comparison
        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1

        result = []
        for i, r in enumerate(rows):
            prev_sales = db.query(func.sum(Sale.amount)).filter(
                Sale.doctor_id == r.id,
                extract('month', Sale.sale_date) == prev_month,
                extract('year', Sale.sale_date) == prev_year
            ).scalar() or 0
            result.append({
                "posicion": i + 1,
                "doctor_id": r.id,
                "nombre": r.name,
                "especialidad": r.specialty,
                "monto_mes": round(float(r.total_monto or 0), 2),
                "unidades": r.total_registros,
                "registros": r.total_registros,
                "monto_mes_anterior": round(float(prev_sales), 2),
                "variacion_pct": round((float(r.total_monto or 0) - float(prev_sales)) / float(prev_sales) * 100, 1) if prev_sales else None
            })

        return {"periodo": f"{month:02d}/{year}", "ranking": result}

    # ── get_new_doctors ───────────────────────────────────────────────────────
    elif tool_name == "get_new_doctors":
        month = tool_input.get("month", current_month)
        year = tool_input.get("year", current_year)

        # First sale ever for each doctor that happened in this month/year
        subq = db.query(
            Sale.doctor_id,
            func.min(Sale.sale_date).label("primera_venta")
        ).group_by(Sale.doctor_id).subquery()

        rows = db.query(
            Doctor.id, Doctor.name, Doctor.specialty, subq.c.primera_venta,
            MedicalRep.name.label("rep_name")
        ).join(subq, subq.c.doctor_id == Doctor.id
        ).outerjoin(MedicalRep, MedicalRep.id == Doctor.rep_id
        ).filter(
            extract('month', subq.c.primera_venta) == month,
            extract('year', subq.c.primera_venta) == year
        ).all()

        result = []
        for r in rows:
            total = db.query(func.sum(Sale.amount)).filter(
                Sale.doctor_id == r.id,
                extract('month', Sale.sale_date) == month,
                extract('year', Sale.sale_date) == year
            ).scalar() or 0
            result.append({
                "doctor_id": r.id,
                "nombre": r.name,
                "especialidad": r.specialty,
                "visitador": r.rep_name,
                "primera_venta": r.primera_venta.strftime("%d/%m/%Y") if r.primera_venta else None,
                "ventas_mes": round(float(total), 2)
            })

        return {"periodo": f"{month:02d}/{year}", "medicos_nuevos": result, "total": len(result)}

    # ── get_sales_analysis ────────────────────────────────────────────────────
    elif tool_name == "get_sales_analysis":
        month = tool_input.get("month", current_month)
        year = tool_input.get("year", current_year)
        rep_id = tool_input.get("rep_id")
        doctor_id = tool_input.get("doctor_id")
        business_line_name = tool_input.get("business_line_name")
        group_by = tool_input.get("group_by", "doctor")

        base_query = db.query(Sale).filter(
            extract('month', Sale.sale_date) == month,
            extract('year', Sale.sale_date) == year
        )

        if doctor_id:
            base_query = base_query.filter(Sale.doctor_id == doctor_id)
        if rep_id:
            doc_ids = [d.id for d in db.query(Doctor.id).filter(Doctor.rep_id == rep_id).all()]
            base_query = base_query.filter(Sale.doctor_id.in_(doc_ids))
        if business_line_name:
            bl = db.query(BusinessLine).filter(BusinessLine.name.ilike(f"%{business_line_name}%")).first()
            if bl:
                doc_ids = [d.id for d in db.query(Doctor.id).filter(Doctor.business_line_id == bl.id).all()]
                base_query = base_query.filter(Sale.doctor_id.in_(doc_ids))

        sales = base_query.all()
        total_monto = sum(float(s.amount or 0) for s in sales)
        total_registros = len(sales)

        grouped = {}
        for s in sales:
            if group_by == "doctor":
                key = s.doctor_id
                label = s.doctor.name if s.doctor else f"Doctor {s.doctor_id}"
            elif group_by == "rep":
                key = s.doctor.rep_id if s.doctor else None
                label = s.doctor.rep.name if s.doctor and s.doctor.rep else "Sin visitador"
            elif group_by == "business_line":
                key = s.doctor.business_line_id if s.doctor else None
                label = s.doctor.business_line.name if s.doctor and s.doctor.business_line else "Sin línea"
            elif group_by == "category":
                key = s.categoria or "Sin categoría"
                label = key
            else:
                key = label = "total"

            if key not in grouped:
                grouped[key] = {"label": label, "monto": 0, "registros": 0}
            grouped[key]["monto"] += float(s.amount or 0)
            grouped[key]["registros"] += 1

        breakdown = sorted(
            [{"grupo": v["label"], "monto": round(v["monto"], 2), "registros": v["registros"]} for v in grouped.values()],
            key=lambda x: x["monto"], reverse=True
        )

        return {
            "periodo": f"{month:02d}/{year}",
            "total_monto": round(total_monto, 2),
            "total_registros": total_registros,
            "agrupado_por": group_by,
            "detalle": breakdown[:50]
        }

    # ── get_rep_commissions ───────────────────────────────────────────────────
    elif tool_name == "get_rep_commissions":
        month = tool_input.get("month", current_month)
        year = tool_input.get("year", current_year)
        filter_rep_id = tool_input.get("rep_id")

        reps_q = db.query(MedicalRep).filter(MedicalRep.is_active == True)
        if filter_rep_id:
            reps_q = reps_q.filter(MedicalRep.id == filter_rep_id)
        reps = reps_q.all()

        result = []
        for rep in reps:
            doctors = db.query(Doctor).filter(Doctor.rep_id == rep.id, Doctor.is_active == True).all()

            # New doctors this month
            new_docs_count = 0
            for doc in doctors:
                first_sale = db.query(Sale).filter(Sale.doctor_id == doc.id).order_by(Sale.sale_date.asc()).first()
                if first_sale and first_sale.sale_date:
                    if first_sale.sale_date.month == month and first_sale.sale_date.year == year:
                        new_docs_count += 1

            # Sales by category
            sales_cats = db.query(
                Sale.categoria, func.sum(Sale.amount).label("total"), func.count(Sale.id).label("count")
            ).join(Doctor, Doctor.id == Sale.doctor_id
            ).filter(
                Doctor.rep_id == rep.id,
                extract('month', Sale.sale_date) == month,
                extract('year', Sale.sale_date) == year
            ).group_by(Sale.categoria).all()

            total_sales = sum(float(r.total or 0) for r in sales_cats)

            # Active doctors (with at least 1 sale this month)
            active_docs = db.query(func.count(func.distinct(Sale.doctor_id))
            ).join(Doctor, Doctor.id == Sale.doctor_id
            ).filter(
                Doctor.rep_id == rep.id,
                extract('month', Sale.sale_date) == month,
                extract('year', Sale.sale_date) == year
            ).scalar()

            result.append({
                "rep_id": rep.id,
                "nombre": rep.name,
                "zona": rep.zone,
                "medicos_activos": active_docs,
                "medicos_nuevos": new_docs_count,
                "total_ventas": round(total_sales, 2),
                "ventas_total": round(total_sales, 2),
                "ventas_por_categoria": [
                    {"categoria": r.categoria or "Sin categoría", "monto": round(float(r.total or 0), 2), "registros": r.count}
                    for r in sorted(sales_cats, key=lambda x: float(x.total or 0), reverse=True)
                ]
            })

        return {"periodo": f"{month:02d}/{year}", "comisiones": result}

    # ── get_visits_tracking ───────────────────────────────────────────────────
    elif tool_name == "get_visits_tracking":
        month = tool_input.get("month")
        year = tool_input.get("year", current_year)
        date_str = tool_input.get("date")

        if month:
            start = datetime(year, month, 1)
            end = datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1)
            label = f"{month:02d}/{year}"
        else:
            if date_str:
                try:
                    d = datetime.strptime(date_str, "%Y-%m-%d")
                except ValueError:
                    d = now
            else:
                d = now
            start = d.replace(hour=0, minute=0, second=0, microsecond=0)
            end = start + timedelta(days=1)
            label = d.strftime("%d/%m/%Y")

        reps = db.query(MedicalRep).filter(MedicalRep.is_active == True).all()
        result = []
        for rep in reps:
            total = db.query(func.count(Visit.id)).filter(
                Visit.rep_id == rep.id, Visit.scheduled_date >= start, Visit.scheduled_date < end
            ).scalar()
            completed = db.query(func.count(Visit.id)).filter(
                Visit.rep_id == rep.id, Visit.scheduled_date >= start, Visit.scheduled_date < end,
                Visit.status == "completed"
            ).scalar()
            missed = db.query(func.count(Visit.id)).filter(
                Visit.rep_id == rep.id, Visit.scheduled_date >= start, Visit.scheduled_date < end,
                Visit.status == "missed"
            ).scalar()
            result.append({
                "rep_id": rep.id,
                "rep_name": rep.name,
                "nombre": rep.name,
                "visitas_programadas": total,
                "completadas": completed,
                "perdidas": missed,
                "completion_rate": round(completed / total * 100, 1) if total else 0,
                "tasa": round(completed / total * 100, 1) if total else 0
            })

        result.sort(key=lambda x: x["tasa"])
        return {"periodo": label, "tracking": result}

    # ── search_doctors ────────────────────────────────────────────────────────
    elif tool_name == "search_doctors":
        search = tool_input.get("search", "")
        specialty = tool_input.get("specialty")
        rep_id = tool_input.get("rep_id")
        without_rep = tool_input.get("without_rep", False)
        limit = tool_input.get("limit", 30)

        q = db.query(Doctor).filter(Doctor.is_active == True)
        if search:
            q = q.filter(Doctor.name.ilike(f"%{search}%"))
        if specialty:
            q = q.filter(Doctor.specialty.ilike(f"%{specialty}%"))
        if rep_id:
            q = q.filter(Doctor.rep_id == rep_id)
        if without_rep:
            q = q.filter(Doctor.rep_id == None)

        doctors = q.limit(limit).all()
        result = []
        for d in doctors:
            last_visit = db.query(Visit).filter(
                Visit.doctor_id == d.id, Visit.status == "completed"
            ).order_by(Visit.actual_date.desc()).first()

            last_sale = db.query(Sale).filter(Sale.doctor_id == d.id).order_by(Sale.sale_date.desc()).first()

            result.append({
                "doctor_id": d.id,
                "nombre": d.name,
                "especialidad": d.specialty,
                "direccion": d.address,
                "telefono": d.phone,
                "visitador": d.rep.name if d.rep else None,
                "linea_negocio": d.business_line.name if d.business_line else None,
                "ultima_visita": last_visit.actual_date.strftime("%d/%m/%Y") if last_visit and last_visit.actual_date else None,
                "ultima_venta": last_sale.sale_date.strftime("%d/%m/%Y") if last_sale and last_sale.sale_date else None
            })

        return {"medicos": result, "total": len(result)}

    # ── get_doctor_detail ─────────────────────────────────────────────────────
    elif tool_name == "get_doctor_detail":
        doctor_id = tool_input.get("doctor_id")
        doctor = db.query(Doctor).filter(Doctor.id == doctor_id).first()
        if not doctor:
            return {"error": f"Médico {doctor_id} no encontrado"}

        # Last 6 visits
        recent_visits = db.query(Visit).filter(Visit.doctor_id == doctor_id
                                               ).order_by(Visit.scheduled_date.desc()).limit(10).all()

        # Sales last 6 months
        six_months_ago = now - timedelta(days=180)
        monthly_sales = db.query(
            extract('year', Sale.sale_date).label("year"),
            extract('month', Sale.sale_date).label("month"),
            func.sum(Sale.amount).label("total"),
            func.count(Sale.id).label("count")
        ).filter(
            Sale.doctor_id == doctor_id,
            Sale.sale_date >= six_months_ago
        ).group_by("year", "month").order_by("year", "month").all()

        total_lifetime = db.query(func.sum(Sale.amount)).filter(Sale.doctor_id == doctor_id).scalar() or 0

        return {
            "doctor_id": doctor.id,
            "nombre": doctor.name,
            "especialidad": doctor.specialty,
            "rut": doctor.rut,
            "direccion": doctor.address,
            "telefono": doctor.phone,
            "email": doctor.email,
            "visitador": doctor.rep.name if doctor.rep else None,
            "linea_negocio": doctor.business_line.name if doctor.business_line else None,
            "frecuencia_visita_dias": doctor.visit_frequency,
            "notas": doctor.notes,
            "ventas_historicas_total": round(float(total_lifetime), 2),
            "ventas_ultimos_6_meses": [
                {"mes": f"{int(r.month):02d}/{int(r.year)}", "monto": round(float(r.total or 0), 2), "registros": r.count}
                for r in monthly_sales
            ],
            "visitas_recientes": [
                {"fecha": v.scheduled_date.strftime("%d/%m/%Y") if v.scheduled_date else None,
                 "estado": v.status, "notas": v.notes}
                for v in recent_visits
            ]
        }

    # ── get_inactive_doctors ──────────────────────────────────────────────────
    elif tool_name == "get_inactive_doctors":
        days = tool_input.get("days_without_visit", 45)
        rep_id = tool_input.get("rep_id")
        cutoff = now - timedelta(days=days)

        q = db.query(Doctor).filter(Doctor.is_active == True)
        if rep_id:
            q = q.filter(Doctor.rep_id == rep_id)
        doctors = q.all()

        inactive = []
        for d in doctors:
            last_visit = db.query(Visit).filter(
                Visit.doctor_id == d.id, Visit.status == "completed"
            ).order_by(Visit.actual_date.desc()).first()

            last_date = last_visit.actual_date if last_visit and last_visit.actual_date else None
            if last_date is None or last_date < cutoff:
                days_since = (now - last_date).days if last_date else None
                inactive.append({
                    "doctor_id": d.id,
                    "nombre": d.name,
                    "especialidad": d.specialty,
                    "visitador": d.rep.name if d.rep else "Sin asignar",
                    "ultima_visita": last_date.strftime("%d/%m/%Y") if last_date else "Nunca visitado",
                    "dias_sin_visita": days_since
                })

        inactive.sort(key=lambda x: (x["dias_sin_visita"] is None, x["dias_sin_visita"] or 999), reverse=True)
        return {"medicos_inactivos": inactive, "total": len(inactive), "criterio_dias": days}

    # ── list_reps ─────────────────────────────────────────────────────────────
    elif tool_name == "list_reps":
        reps = db.query(MedicalRep).filter(MedicalRep.is_active == True).all()
        return {
            "visitadores": [
                {"rep_id": r.id, "nombre": r.name, "email": r.email,
                 "telefono": r.phone, "zona": r.zone, "territorio": r.territory}
                for r in reps
            ]
        }

    # ── assign_rep_to_doctor ──────────────────────────────────────────────────
    elif tool_name == "assign_rep_to_doctor":
        doctor_id = tool_input.get("doctor_id")
        rep_id = tool_input.get("rep_id")

        doctor = db.query(Doctor).filter(Doctor.id == doctor_id).first()
        if not doctor:
            return {"error": f"Médico {doctor_id} no encontrado"}

        old_rep = doctor.rep.name if doctor.rep else "Sin asignar"

        if rep_id == 0:
            doctor.rep_id = None
            new_rep = "Sin asignar"
        else:
            rep = db.query(MedicalRep).filter(MedicalRep.id == rep_id).first()
            if not rep:
                return {"error": f"Visitador {rep_id} no encontrado"}
            doctor.rep_id = rep_id
            new_rep = rep.name

        db.commit()
        return {
            "success": True,
            "doctor": doctor.name,
            "visitador_anterior": old_rep,
            "visitador_nuevo": new_rep
        }

    # ── get_monthly_trend ─────────────────────────────────────────────────────
    elif tool_name == "get_monthly_trend":
        months_back = tool_input.get("months", 6)
        rep_id = tool_input.get("rep_id")
        business_line_name = tool_input.get("business_line_name")

        bl_id = None
        if business_line_name:
            bl = db.query(BusinessLine).filter(BusinessLine.name.ilike(f"%{business_line_name}%")).first()
            if bl:
                bl_id = bl.id

        result = []
        for i in range(months_back - 1, -1, -1):
            m = current_month - i
            y = current_year
            while m <= 0:
                m += 12
                y -= 1

            month_start = datetime(y, m, 1)
            month_end = datetime(y, m + 1, 1) if m < 12 else datetime(y + 1, 1, 1)

            sales_q = db.query(func.sum(Sale.amount), func.count(Sale.id)
                               ).join(Doctor, Doctor.id == Sale.doctor_id
                               ).filter(
                extract('month', Sale.sale_date) == m,
                extract('year', Sale.sale_date) == y
            )
            if rep_id:
                sales_q = sales_q.filter(Doctor.rep_id == rep_id)
            if bl_id:
                sales_q = sales_q.filter(Doctor.business_line_id == bl_id)

            monto, count = sales_q.first()

            visits_q = db.query(func.count(Visit.id)).filter(
                Visit.scheduled_date >= month_start,
                Visit.scheduled_date < month_end,
                Visit.status == "completed"
            )
            if rep_id:
                visits_q = visits_q.filter(Visit.rep_id == rep_id)
            visitas = visits_q.scalar()

            result.append({
                "mes": f"{m:02d}/{y}",
                "ventas_monto": round(float(monto or 0), 2),
                "ventas_registros": count or 0,
                "visitas_completadas": visitas
            })

        return {"tendencia": result, "meses_analizados": months_back}

    # ── export_to_excel ───────────────────────────────────────────────────────
    elif tool_name == "export_to_excel":
        data_type = tool_input.get("data_type")
        month = tool_input.get("month", current_month)
        year = tool_input.get("year", current_year)
        rep_id = tool_input.get("rep_id")

        wb = openpyxl.Workbook()
        ws = wb.active

        # Header style
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="7C3AED")
        header_align = Alignment(horizontal="center")

        def style_headers(ws, headers):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

        rows_written = 0

        if data_type == "ranking_medicos":
            ws.title = "Ranking Médicos"
            data = execute_mike_tool("get_doctor_ranking", {"month": month, "year": year, "top": 50, **({"rep_id": rep_id} if rep_id else {})}, db)
            headers = ["Posición", "Médico", "Especialidad", "Monto Mes", "Registros", "Monto Mes Anterior", "Variación %"]
            style_headers(ws, headers)
            for row in data.get("ranking", []):
                ws.append([
                    row.get("posicion"), row.get("nombre"), row.get("especialidad"),
                    row.get("monto_mes"), row.get("registros"),
                    row.get("monto_mes_anterior"), row.get("variacion_pct")
                ])
                rows_written += 1

        elif data_type == "visitadores":
            ws.title = "Visitadores"
            data = execute_mike_tool("get_all_reps_summary", {"month": month, "year": year}, db)
            headers = ["Visitador", "Zona", "Territorio", "Médicos Asignados", "Visitas Programadas", "Visitas Completadas", "Tasa Cumplimiento %", "Ventas Monto"]
            style_headers(ws, headers)
            for row in data.get("visitadores", []):
                ws.append([
                    row.get("nombre"), row.get("zona"), row.get("territorio"),
                    row.get("medicos_asignados"), row.get("visitas_programadas"),
                    row.get("visitas_completadas"), row.get("tasa_cumplimiento"),
                    row.get("ventas_monto")
                ])
                rows_written += 1

        elif data_type == "nuevos_medicos":
            ws.title = "Médicos Nuevos"
            data = execute_mike_tool("get_new_doctors", {"month": month, "year": year}, db)
            headers = ["Médico", "Especialidad", "Visitador", "Primera Venta", "Ventas Mes"]
            style_headers(ws, headers)
            for row in data.get("medicos_nuevos", []):
                ws.append([
                    row.get("nombre"), row.get("especialidad"), row.get("visitador"),
                    row.get("primera_venta"), row.get("ventas_mes")
                ])
                rows_written += 1

        elif data_type == "tendencia_ventas":
            ws.title = "Tendencia Ventas"
            data = execute_mike_tool("get_monthly_trend", {"months": 12, **({"rep_id": rep_id} if rep_id else {})}, db)
            headers = ["Mes", "Ventas Monto", "Ventas Registros", "Visitas Completadas"]
            style_headers(ws, headers)
            for row in data.get("tendencia", []):
                ws.append([
                    row.get("mes"), row.get("ventas_monto"),
                    row.get("ventas_registros"), row.get("visitas_completadas")
                ])
                rows_written += 1

        elif data_type == "comisiones":
            ws.title = "Comisiones"
            data = execute_mike_tool("get_rep_commissions", {"month": month, "year": year, **({"rep_id": rep_id} if rep_id else {})}, db)
            headers = ["Visitador", "Zona", "Médicos Activos", "Médicos Nuevos", "Ventas Total"]
            style_headers(ws, headers)
            for row in data.get("comisiones", []):
                ws.append([
                    row.get("nombre"), row.get("zona"),
                    row.get("medicos_activos"), row.get("medicos_nuevos"),
                    row.get("ventas_total")
                ])
                rows_written += 1

        elif data_type == "ranking_pacientes":
            ws.title = "Ranking Pacientes"
            data = execute_mike_tool("get_patient_analysis", {"month": month, "year": year, "top": 100, "analysis_type": "ranking"}, db)
            headers = ["Posición", "RUT", "Nombre", "Total Monto", "N° Compras", "Última Compra", "Médico"]
            style_headers(ws, headers)
            for row in data.get("ranking", []):
                ws.append([
                    row.get("posicion"), row.get("rut"), row.get("nombre"),
                    row.get("total_monto"), row.get("total_compras"),
                    row.get("ultima_compra"), row.get("medico")
                ])
                rows_written += 1

        elif data_type == "pacientes_nuevos":
            ws.title = "Pacientes Nuevos"
            data = execute_mike_tool("get_patient_analysis", {"month": month, "year": year, "analysis_type": "new_patients"}, db)
            headers = ["RUT", "Nombre", "Primera Compra", "Médico", "Visitador", "Total Mes"]
            style_headers(ws, headers)
            for row in data.get("pacientes_nuevos", []):
                ws.append([
                    row.get("rut"), row.get("nombre"), row.get("primera_compra"),
                    row.get("medico"), row.get("visitador"), row.get("total_mes")
                ])
                rows_written += 1

        # Auto-size columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Save to BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        token = uuid.uuid4().hex[:8]
        filename = f"mike_export_{data_type}_{month:02d}{year}.xlsx"
        _export_store[token] = {"data": buffer.getvalue(), "filename": filename}

        return {
            "download_ready": True,
            "token": token,
            "filename": filename,
            "rows": rows_written
        }

    # ── get_rep_calendar ──────────────────────────────────────────────────────
    elif tool_name == "get_rep_calendar":
        rep_id    = tool_input.get("rep_id")
        date_from = tool_input.get("date_from", now.strftime("%Y-%m-%d"))
        date_to   = tool_input.get("date_to", (now + timedelta(days=30)).strftime("%Y-%m-%d"))
        status    = tool_input.get("status")

        rep = db.query(MedicalRep).filter(MedicalRep.id == rep_id).first()
        if not rep:
            return {"error": f"Visitador {rep_id} no encontrado"}

        q = db.query(Visit).filter(
            Visit.rep_id == rep_id,
            Visit.scheduled_date >= datetime.strptime(date_from, "%Y-%m-%d"),
            Visit.scheduled_date < datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
        )
        if status:
            q = q.filter(Visit.status == status)
        visits = q.order_by(Visit.scheduled_date).all()

        by_day: dict = {}
        for v in visits:
            day_key = v.scheduled_date.strftime("%Y-%m-%d") if v.scheduled_date else "sin fecha"
            doctor  = db.query(Doctor).filter(Doctor.id == v.doctor_id).first()
            by_day.setdefault(day_key, []).append({
                "visit_id":   v.id,
                "medico":     doctor.name if doctor else v.doctor_id,
                "especialidad": doctor.specialty if doctor else None,
                "centro":     doctor.medical_center if doctor else None,
                "ciudad":     doctor.city if doctor else None,
                "telefono":   doctor.phone if doctor else None,
                "estado":     v.status,
                "notas":      v.notes
            })

        return {
            "visitador": rep.name,
            "periodo":   f"{date_from} → {date_to}",
            "total_visitas": len(visits),
            "por_dia": by_day
        }

    # ── schedule_visit ────────────────────────────────────────────────────────
    elif tool_name == "schedule_visit":
        rep_id    = tool_input.get("rep_id")
        doctor_id = tool_input.get("doctor_id")
        date_str  = tool_input.get("scheduled_date")
        notes     = tool_input.get("notes")

        rep    = db.query(MedicalRep).filter(MedicalRep.id == rep_id).first()
        doctor = db.query(Doctor).filter(Doctor.id == doctor_id).first()
        if not rep:    return {"error": f"Visitador {rep_id} no encontrado"}
        if not doctor: return {"error": f"Médico {doctor_id} no encontrado"}

        visit = Visit(
            rep_id=rep_id,
            doctor_id=doctor_id,
            scheduled_date=datetime.strptime(date_str, "%Y-%m-%d"),
            status="scheduled",
            notes=notes
        )
        db.add(visit)
        db.commit()
        db.refresh(visit)
        return {
            "success": True,
            "visit_id": visit.id,
            "visitador": rep.name,
            "medico": doctor.name,
            "fecha": date_str
        }

    # ── reschedule_visit ──────────────────────────────────────────────────────
    elif tool_name == "reschedule_visit":
        visit_id = tool_input.get("visit_id")
        new_date  = tool_input.get("new_date")
        notes     = tool_input.get("notes")

        visit = db.query(Visit).filter(Visit.id == visit_id).first()
        if not visit: return {"error": f"Visita {visit_id} no encontrada"}

        old_date = visit.scheduled_date.strftime("%Y-%m-%d") if visit.scheduled_date else "?"
        visit.scheduled_date = datetime.strptime(new_date, "%Y-%m-%d")
        if notes: visit.notes = notes
        db.commit()

        doctor = db.query(Doctor).filter(Doctor.id == visit.doctor_id).first()
        return {
            "success": True,
            "visit_id": visit_id,
            "medico": doctor.name if doctor else visit.doctor_id,
            "fecha_anterior": old_date,
            "fecha_nueva": new_date
        }

    # ── cancel_visit ──────────────────────────────────────────────────────────
    elif tool_name == "cancel_visit":
        visit_id = tool_input.get("visit_id")
        reason   = tool_input.get("reason", "")

        visit = db.query(Visit).filter(Visit.id == visit_id).first()
        if not visit: return {"error": f"Visita {visit_id} no encontrada"}

        visit.status = "cancelled"
        if reason: visit.notes = f"[Cancelada: {reason}] {visit.notes or ''}".strip()
        db.commit()

        doctor = db.query(Doctor).filter(Doctor.id == visit.doctor_id).first()
        return {
            "success": True,
            "visit_id": visit_id,
            "medico": doctor.name if doctor else visit.doctor_id,
            "fecha": visit.scheduled_date.strftime("%Y-%m-%d") if visit.scheduled_date else "?"
        }

    # ── bulk_assign_doctors ───────────────────────────────────────────────────
    elif tool_name == "bulk_assign_doctors":
        rep_id          = tool_input.get("rep_id")
        business_line_name = tool_input.get("business_line_name")
        specialty       = tool_input.get("specialty")
        city            = tool_input.get("city")
        without_rep_only = tool_input.get("without_rep_only", True)
        limit           = tool_input.get("limit", 50)
        doctor_ids      = tool_input.get("doctor_ids")

        rep = db.query(MedicalRep).filter(MedicalRep.id == rep_id).first()
        if not rep: return {"error": f"Visitador {rep_id} no encontrado"}

        if doctor_ids:
            doctors = db.query(Doctor).filter(Doctor.id.in_(doctor_ids), Doctor.is_active == True).all()
        else:
            q = db.query(Doctor).filter(Doctor.is_active == True)
            if without_rep_only:
                q = q.filter(Doctor.rep_id == None)
            if business_line_name:
                bl = db.query(BusinessLine).filter(BusinessLine.name.ilike(f"%{business_line_name}%")).first()
                if bl: q = q.filter(Doctor.business_line_id == bl.id)
            if specialty:
                q = q.filter(Doctor.specialty.ilike(f"%{specialty}%"))
            if city:
                q = q.filter(Doctor.city.ilike(f"%{city}%"))
            doctors = q.limit(limit).all()

        assigned = []
        for doc in doctors:
            doc.rep_id = rep_id
            assigned.append({"doctor_id": doc.id, "nombre": doc.name, "ciudad": doc.city, "especialidad": doc.specialty})
        db.commit()

        return {
            "success": True,
            "visitador": rep.name,
            "medicos_asignados": len(assigned),
            "detalle": assigned
        }

    # ── get_rep_behavior_analysis ─────────────────────────────────────────────
    elif tool_name == "get_rep_behavior_analysis":
        rep_id     = tool_input.get("rep_id")
        months_back = tool_input.get("months", 3)

        rep = db.query(MedicalRep).filter(MedicalRep.id == rep_id).first()
        if not rep: return {"error": f"Visitador {rep_id} no encontrado"}

        since = now - timedelta(days=30 * months_back)
        visits = db.query(Visit).filter(
            Visit.rep_id == rep_id,
            Visit.scheduled_date >= since
        ).order_by(Visit.scheduled_date).all()

        # By status
        total = len(visits)
        completed = [v for v in visits if v.status == "completed"]
        missed    = [v for v in visits if v.status == "missed"]
        scheduled = [v for v in visits if v.status == "scheduled"]

        # Activity by weekday
        weekday_names = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
        by_weekday = {d: {"programadas": 0, "completadas": 0} for d in weekday_names}
        for v in visits:
            if v.scheduled_date:
                day = weekday_names[v.scheduled_date.weekday()]
                by_weekday[day]["programadas"] += 1
                if v.status == "completed":
                    by_weekday[day]["completadas"] += 1

        # By month
        by_month: dict = {}
        for v in visits:
            if v.scheduled_date:
                key = v.scheduled_date.strftime("%Y-%m")
                by_month.setdefault(key, {"programadas": 0, "completadas": 0, "perdidas": 0})
                by_month[key]["programadas"] += 1
                if v.status == "completed": by_month[key]["completadas"] += 1
                if v.status == "missed":    by_month[key]["perdidas"] += 1

        for key in by_month:
            prog = by_month[key]["programadas"]
            comp = by_month[key]["completadas"]
            by_month[key]["tasa"] = round(comp / prog * 100, 1) if prog else 0

        # Doctors visited vs assigned
        doctors_assigned = db.query(func.count(Doctor.id)).filter(
            Doctor.rep_id == rep_id, Doctor.is_active == True
        ).scalar()
        doctors_visited_ids = {v.doctor_id for v in completed}

        # Sales correlation
        sales_by_month: dict = {}
        for i in range(months_back):
            m = current_month - i
            y = current_year
            while m <= 0: m += 12; y -= 1
            amount = db.query(func.sum(Sale.amount)).join(
                Doctor, Doctor.id == Sale.doctor_id
            ).filter(
                Doctor.rep_id == rep_id,
                extract('month', Sale.sale_date) == m,
                extract('year',  Sale.sale_date) == y
            ).scalar() or 0
            sales_by_month[f"{y}-{m:02d}"] = round(float(amount), 2)

        # Comparison with other reps
        all_reps = db.query(MedicalRep).filter(MedicalRep.is_active == True).all()
        team_rates = []
        for r in all_reps:
            rv = db.query(Visit).filter(Visit.rep_id == r.id, Visit.scheduled_date >= since).all()
            rc = [v for v in rv if v.status == "completed"]
            team_rates.append({
                "rep_id": r.id,
                "nombre": r.name,
                "tasa": round(len(rc) / len(rv) * 100, 1) if rv else 0,
                "visitas_totales": len(rv)
            })
        team_avg = round(sum(r["tasa"] for r in team_rates) / len(team_rates), 1) if team_rates else 0
        my_rate  = round(len(completed) / total * 100, 1) if total else 0

        return {
            "visitador": rep.name,
            "periodo_analizado": f"Últimos {months_back} meses",
            "resumen": {
                "total_visitas_programadas": total,
                "visitas_completadas": len(completed),
                "visitas_perdidas": len(missed),
                "visitas_pendientes": len(scheduled),
                "tasa_cumplimiento": my_rate,
                "promedio_equipo": team_avg,
                "vs_equipo": f"{my_rate - team_avg:+.1f}% vs promedio"
            },
            "medicos": {
                "asignados": doctors_assigned,
                "visitados_periodo": len(doctors_visited_ids),
                "cobertura": round(len(doctors_visited_ids) / doctors_assigned * 100, 1) if doctors_assigned else 0
            },
            "actividad_por_dia": by_weekday,
            "tendencia_mensual": by_month,
            "ventas_por_mes": sales_by_month,
            "comparacion_equipo": sorted(team_rates, key=lambda x: x["tasa"], reverse=True)
        }

    # ── save_to_memory ────────────────────────────────────────────────────────
    elif tool_name == "save_to_memory":
        content  = tool_input.get("content", "").strip()
        category = tool_input.get("category", "general")
        if not content:
            return {"error": "El contenido no puede estar vacío"}
        mem = MikeMemory(content=content, category=category)
        db.add(mem)
        db.commit()
        db.refresh(mem)
        return {"success": True, "memory_id": mem.id, "content": mem.content, "category": mem.category}

    # ── get_memories ──────────────────────────────────────────────────────────
    elif tool_name == "get_memories":
        category = tool_input.get("category")
        q = db.query(MikeMemory)
        if category:
            q = q.filter(MikeMemory.category == category)
        mems = q.order_by(MikeMemory.created_at.desc()).all()
        return {
            "total": len(mems),
            "memories": [
                {"id": m.id, "content": m.content, "category": m.category,
                 "fecha": m.created_at.strftime("%d/%m/%Y") if m.created_at else ""}
                for m in mems
            ]
        }

    # ── delete_memory ─────────────────────────────────────────────────────────
    elif tool_name == "delete_memory":
        mem_id = tool_input.get("memory_id")
        mem = db.query(MikeMemory).filter(MikeMemory.id == mem_id).first()
        if not mem:
            return {"error": f"Recuerdo {mem_id} no encontrado"}
        db.delete(mem)
        db.commit()
        return {"success": True, "deleted_id": mem_id}

    # ── assign_business_line_to_doctor ────────────────────────────────────────
    elif tool_name == "assign_business_line_to_doctor":
        doctor_id = tool_input.get("doctor_id")
        bl_name   = tool_input.get("business_line_name", "")

        doctor = db.query(Doctor).filter(Doctor.id == doctor_id).first()
        if not doctor:
            return {"error": f"Médico {doctor_id} no encontrado"}

        bl = db.query(BusinessLine).filter(BusinessLine.name.ilike(f"%{bl_name}%")).first()
        if not bl:
            all_bls = [b.name for b in db.query(BusinessLine).all()]
            return {"error": f"Línea '{bl_name}' no encontrada. Disponibles: {', '.join(all_bls)}"}

        old_bl = doctor.business_line.name if doctor.business_line else "Sin línea"
        doctor.business_line_id = bl.id
        db.commit()

        return {
            "success": True,
            "doctor": doctor.name,
            "linea_anterior": old_bl,
            "linea_nueva": bl.name
        }

    # ── auto_assign_business_lines ────────────────────────────────────────────
    elif tool_name == "auto_assign_business_lines":
        only_without = tool_input.get("only_without_line", True)
        rep_id_filter = tool_input.get("rep_id")
        dry_run = tool_input.get("dry_run", False)

        # Mapeo categoría de venta → nombre de línea de negocio
        CATEGORY_TO_LINE = {
            "hormonas": "Hormonas",
            "dermatología": "Dermatología",
            "dermatologia": "Dermatología",
            "control de peso": "Control de Peso",
            "cannabis medicinal": "Cannabis Medicinal",
            "suero terapia": "Suero Terapia",
            "fertilidad": "Hormonas",
            "pelo": "Dermatología",
            "producto terminado": "Hormonas",
            "veterinaria": "Veterinaria",
        }

        # Precargar líneas de negocio
        all_bls = {bl.name: bl for bl in db.query(BusinessLine).all()}

        # Médicos a procesar
        q = db.query(Doctor).filter(Doctor.is_active == True)
        if only_without:
            q = q.filter(Doctor.business_line_id == None)
        if rep_id_filter:
            q = q.filter(Doctor.rep_id == rep_id_filter)
        doctors = q.all()

        assigned = []
        skipped = []

        for doc in doctors:
            # Contar ventas por categoría para este médico
            cat_counts = db.query(
                Sale.categoria, func.count(Sale.id).label("n")
            ).filter(
                Sale.doctor_id == doc.id,
                Sale.categoria != None
            ).group_by(Sale.categoria).order_by(func.count(Sale.id).desc()).all()

            if not cat_counts:
                skipped.append({"doctor_id": doc.id, "nombre": doc.name, "razon": "sin ventas registradas"})
                continue

            # Categoría más frecuente
            top_cat = cat_counts[0].categoria or ""
            bl_name = CATEGORY_TO_LINE.get(top_cat.lower())

            if not bl_name:
                # Intentar match parcial
                for key, val in CATEGORY_TO_LINE.items():
                    if key in top_cat.lower():
                        bl_name = val
                        break

            if not bl_name or bl_name not in all_bls:
                skipped.append({"doctor_id": doc.id, "nombre": doc.name, "razon": f"categoría '{top_cat}' sin mapeo"})
                continue

            bl = all_bls[bl_name]
            old_bl = doc.business_line.name if doc.business_line else "Sin línea"

            if not dry_run:
                doc.business_line_id = bl.id

            assigned.append({
                "doctor_id": doc.id,
                "nombre": doc.name,
                "linea_anterior": old_bl,
                "linea_nueva": bl_name,
                "categoria_principal": top_cat
            })

        if not dry_run:
            db.commit()

        return {
            "modo": "preview (sin cambios)" if dry_run else "aplicado",
            "asignados": len(assigned),
            "omitidos": len(skipped),
            "detalle_asignados": assigned[:50],
            "detalle_omitidos": skipped[:20]
        }

    # ── get_patient_analysis ──────────────────────────────────────────────────
    elif tool_name == "get_patient_analysis":
        month = tool_input.get("month", current_month)
        year = tool_input.get("year", current_year)
        top = tool_input.get("top", 20)
        analysis_type = tool_input.get("analysis_type", "ranking")
        days_threshold = tool_input.get("days_without_purchase", 60)
        doctor_id_filter = tool_input.get("doctor_id")
        rep_id_filter = tool_input.get("rep_id")
        categoria_filter = tool_input.get("categoria")

        # Base filter for rep/doctor
        def base_patient_query(db):
            q = db.query(Sale).filter(Sale.rut_paciente != None, Sale.rut_paciente != "")
            if doctor_id_filter:
                q = q.filter(Sale.doctor_id == doctor_id_filter)
            if rep_id_filter:
                doc_ids = [d.id for d in db.query(Doctor.id).filter(Doctor.rep_id == rep_id_filter).all()]
                q = q.filter(Sale.doctor_id.in_(doc_ids))
            if categoria_filter:
                q = q.filter(Sale.categoria.ilike(f"%{categoria_filter}%"))
            return q

        if analysis_type == "ranking":
            # Top patients by purchase volume this month
            rows = db.query(
                Sale.rut_paciente,
                Sale.nombre_paciente,
                func.sum(Sale.amount).label("total"),
                func.count(Sale.id).label("count"),
                func.max(Sale.sale_date).label("ultima_compra")
            ).filter(
                Sale.rut_paciente != None, Sale.rut_paciente != "",
                extract('month', Sale.sale_date) == month,
                extract('year', Sale.sale_date) == year
            )
            if doctor_id_filter:
                rows = rows.filter(Sale.doctor_id == doctor_id_filter)
            if rep_id_filter:
                doc_ids = [d.id for d in db.query(Doctor.id).filter(Doctor.rep_id == rep_id_filter).all()]
                rows = rows.filter(Sale.doctor_id.in_(doc_ids))
            if categoria_filter:
                rows = rows.filter(Sale.categoria.ilike(f"%{categoria_filter}%"))

            rows = rows.group_by(Sale.rut_paciente, Sale.nombre_paciente
                                 ).order_by(func.sum(Sale.amount).desc()).limit(top).all()

            result = []
            for i, r in enumerate(rows):
                # Get doctor info for this patient this month
                doctor_row = db.query(Sale.doctor_id).filter(
                    Sale.rut_paciente == r.rut_paciente,
                    extract('month', Sale.sale_date) == month,
                    extract('year', Sale.sale_date) == year
                ).first()
                doctor = db.query(Doctor).filter(Doctor.id == doctor_row.doctor_id).first() if doctor_row else None
                result.append({
                    "posicion": i + 1,
                    "rut": r.rut_paciente,
                    "nombre": r.nombre_paciente,
                    "total_monto": round(float(r.total or 0), 2),
                    "total_compras": r.count,
                    "ultima_compra": r.ultima_compra.strftime("%d/%m/%Y") if r.ultima_compra else None,
                    "medico": doctor.name if doctor else None
                })
            total_patients = db.query(func.count(func.distinct(Sale.rut_paciente))).filter(
                Sale.rut_paciente != None,
                extract('month', Sale.sale_date) == month,
                extract('year', Sale.sale_date) == year
            ).scalar()
            return {"periodo": f"{month:02d}/{year}", "total_pacientes_mes": total_patients, "ranking": result}

        elif analysis_type == "new_patients":
            # Patients whose first purchase ever is in this month/year
            subq = db.query(
                Sale.rut_paciente,
                func.min(Sale.sale_date).label("primera_compra")
            ).filter(Sale.rut_paciente != None, Sale.rut_paciente != ""
            ).group_by(Sale.rut_paciente).subquery()

            rows = db.query(
                subq.c.rut_paciente,
                Sale.nombre_paciente,
                subq.c.primera_compra,
                Sale.doctor_id
            ).join(Sale, (Sale.rut_paciente == subq.c.rut_paciente) & (Sale.sale_date == subq.c.primera_compra)
            ).filter(
                extract('month', subq.c.primera_compra) == month,
                extract('year', subq.c.primera_compra) == year
            ).all()

            result = []
            for r in rows:
                doctor = db.query(Doctor).filter(Doctor.id == r.doctor_id).first()
                total_mes = db.query(func.sum(Sale.amount)).filter(
                    Sale.rut_paciente == r.rut_paciente,
                    extract('month', Sale.sale_date) == month,
                    extract('year', Sale.sale_date) == year
                ).scalar() or 0
                result.append({
                    "rut": r.rut_paciente,
                    "nombre": r.nombre_paciente,
                    "primera_compra": r.primera_compra.strftime("%d/%m/%Y") if r.primera_compra else None,
                    "medico": doctor.name if doctor else None,
                    "visitador": doctor.rep.name if doctor and doctor.rep else None,
                    "total_mes": round(float(total_mes), 2)
                })
            return {"periodo": f"{month:02d}/{year}", "pacientes_nuevos": result, "total": len(result)}

        elif analysis_type == "retention":
            # Patients who haven't purchased in N days
            cutoff = now - timedelta(days=days_threshold)
            # Get all known patients with their last purchase date
            rows = db.query(
                Sale.rut_paciente,
                Sale.nombre_paciente,
                func.max(Sale.sale_date).label("ultima_compra"),
                func.count(Sale.id).label("total_historico")
            ).filter(Sale.rut_paciente != None, Sale.rut_paciente != ""
            ).group_by(Sale.rut_paciente, Sale.nombre_paciente
            ).having(func.max(Sale.sale_date) < cutoff
            ).order_by(func.max(Sale.sale_date).asc()).limit(top).all()

            result = []
            for r in rows:
                last_doctor_sale = db.query(Sale).filter(
                    Sale.rut_paciente == r.rut_paciente
                ).order_by(Sale.sale_date.desc()).first()
                doctor = db.query(Doctor).filter(Doctor.id == last_doctor_sale.doctor_id).first() if last_doctor_sale else None
                dias_sin_compra = (now - r.ultima_compra).days if r.ultima_compra else None
                result.append({
                    "rut": r.rut_paciente,
                    "nombre": r.nombre_paciente,
                    "ultima_compra": r.ultima_compra.strftime("%d/%m/%Y") if r.ultima_compra else None,
                    "dias_sin_compra": dias_sin_compra,
                    "total_compras_historico": r.total_historico,
                    "ultimo_medico": doctor.name if doctor else None,
                    "visitador": doctor.rep.name if doctor and doctor.rep else None
                })
            return {
                "criterio_dias": days_threshold,
                "pacientes_en_riesgo": result,
                "total": len(result)
            }

        elif analysis_type == "by_category":
            # Distribution by product category this month
            rows = db.query(
                Sale.categoria,
                func.count(func.distinct(Sale.rut_paciente)).label("pacientes_unicos"),
                func.count(Sale.id).label("compras"),
                func.sum(Sale.amount).label("monto")
            ).filter(
                Sale.rut_paciente != None, Sale.rut_paciente != "",
                extract('month', Sale.sale_date) == month,
                extract('year', Sale.sale_date) == year
            ).group_by(Sale.categoria).order_by(func.count(func.distinct(Sale.rut_paciente)).desc()).all()

            return {
                "periodo": f"{month:02d}/{year}",
                "por_categoria": [
                    {
                        "categoria": r.categoria or "Sin categoría",
                        "pacientes_unicos": r.pacientes_unicos,
                        "compras": r.compras,
                        "monto": round(float(r.monto or 0), 2)
                    }
                    for r in rows
                ]
            }

        return {"error": "analysis_type no reconocido"}

    # ── get_patient_detail ────────────────────────────────────────────────────
    elif tool_name == "get_patient_detail":
        rut = tool_input.get("rut_paciente", "")
        nombre = tool_input.get("nombre_paciente", "")

        import re
        rut_norm = re.sub(r'[\.\-\s]', '', rut).upper() if rut else None

        q = db.query(Sale).filter(Sale.rut_paciente != None, Sale.rut_paciente != "")
        if rut_norm:
            # Match normalized RUT
            all_sales_ruts = db.query(Sale.rut_paciente).distinct().all()
            matching_ruts = [
                r[0] for r in all_sales_ruts
                if re.sub(r'[\.\-\s]', '', r[0] or '').upper() == rut_norm
            ]
            if matching_ruts:
                q = q.filter(Sale.rut_paciente.in_(matching_ruts))
            else:
                return {"error": f"No se encontraron compras para RUT '{rut}'"}
        elif nombre:
            q = q.filter(Sale.nombre_paciente.ilike(f"%{nombre}%"))
        else:
            return {"error": "Debes proporcionar rut_paciente o nombre_paciente"}

        sales = q.order_by(Sale.sale_date.desc()).all()
        if not sales:
            return {"error": "No se encontraron compras para este paciente"}

        patient_name = sales[0].nombre_paciente
        patient_rut = sales[0].rut_paciente

        # Monthly trend
        monthly: dict = {}
        for s in sales:
            if s.sale_date:
                key = s.sale_date.strftime("%Y-%m")
                monthly.setdefault(key, {"monto": 0, "compras": 0, "productos": set()})
                monthly[key]["monto"] += float(s.amount or 0)
                monthly[key]["compras"] += 1
                if s.product:
                    monthly[key]["productos"].add(s.product)

        # Doctors attended
        doctor_ids = {s.doctor_id for s in sales if s.doctor_id}
        doctors_info = []
        for did in doctor_ids:
            doc = db.query(Doctor).filter(Doctor.id == did).first()
            if doc:
                doc_sales = [s for s in sales if s.doctor_id == did]
                doctors_info.append({
                    "doctor_id": did,
                    "nombre": doc.name,
                    "especialidad": doc.specialty,
                    "visitador": doc.rep.name if doc.rep else None,
                    "compras": len(doc_sales),
                    "ultima_consulta": max((s.sale_date for s in doc_sales if s.sale_date), default=None)
                })

        # Products summary
        products: dict = {}
        for s in sales:
            cat = s.categoria or "Sin categoría"
            prod = s.product or "Sin producto"
            key = (cat, prod)
            products.setdefault(key, {"count": 0, "monto": 0})
            products[key]["count"] += 1
            products[key]["monto"] += float(s.amount or 0)

        products_list = sorted(
            [{"categoria": k[0], "producto": k[1], "compras": v["count"], "monto": round(v["monto"], 2)}
             for k, v in products.items()],
            key=lambda x: x["compras"], reverse=True
        )

        total_lifetime = sum(float(s.amount or 0) for s in sales)
        monthly_sorted = sorted(
            [{"mes": k, "monto": round(v["monto"], 2), "compras": v["compras"], "productos": list(v["productos"])}
             for k, v in monthly.items()],
            key=lambda x: x["mes"]
        )

        doctors_info.sort(key=lambda x: x["compras"], reverse=True)
        for d in doctors_info:
            if d["ultima_consulta"]:
                d["ultima_consulta"] = d["ultima_consulta"].strftime("%d/%m/%Y")

        return {
            "rut": patient_rut,
            "nombre": patient_name,
            "total_compras_historico": len(sales),
            "gasto_total": round(total_lifetime, 2),
            "primera_compra": min((s.sale_date for s in sales if s.sale_date), default=None).strftime("%d/%m/%Y") if any(s.sale_date for s in sales) else None,
            "ultima_compra": max((s.sale_date for s in sales if s.sale_date), default=None).strftime("%d/%m/%Y") if any(s.sale_date for s in sales) else None,
            "medicos_atendidos": doctors_info,
            "productos": products_list[:20],
            "evolucion_mensual": monthly_sorted
        }

    return {"error": f"Herramienta desconocida: {tool_name}"}


# ── Schemas ───────────────────────────────────────────────────────────────────

from pydantic import BaseModel


class MikeChatRequest(BaseModel):
    message: str
    conversation_history: Optional[List[AgentMessage]] = []


class MikeChatResponse(BaseModel):
    response: str
    conversation_history: List[AgentMessage]
    charts: List[dict] = []
    export_url: Optional[str] = None


# ── Export download endpoint ──────────────────────────────────────────────────

@router.get("/export/{token}")
def download_export(token: str):
    entry = _export_store.get(token)
    if not entry:
        raise HTTPException(status_code=404, detail="Export no encontrado o expirado")

    return StreamingResponse(
        io.BytesIO(entry["data"]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={entry['filename']}"}
    )


# ── Weekly report endpoint ────────────────────────────────────────────────────

@router.post("/weekly-report")
def weekly_report(db: Session = Depends(get_db)):
    now = datetime.utcnow()
    month = now.month
    year = now.year

    # Gather data
    overview = execute_mike_tool("get_dashboard_overview", {}, db)
    reps_summary = execute_mike_tool("get_all_reps_summary", {"month": month, "year": year}, db)
    ranking = execute_mike_tool("get_doctor_ranking", {"month": month, "year": year, "top": 10}, db)

    lines = []
    lines.append(f"REPORTE SEMANAL NARMA — {now.strftime('%d/%m/%Y %H:%M')}")
    lines.append("=" * 60)
    lines.append("")
    lines.append("RESUMEN GENERAL")
    lines.append(f"  Médicos activos: {overview.get('medicos_activos', 0)}")
    lines.append(f"  Visitadores activos: {overview.get('visitadores_activos', 0)}")
    vm = overview.get("visitas_mes", {})
    lines.append(f"  Visitas del mes: {vm.get('completadas', 0)}/{vm.get('programadas', 0)} ({vm.get('tasa_cumplimiento', 0)}%)")
    vn = overview.get("ventas_mes", {})
    lines.append(f"  Ventas del mes: ${vn.get('total_monto', 0):,.2f} ({vn.get('total_registros', 0)} registros)")
    lines.append("")

    lines.append("DESEMPEÑO VISITADORES")
    for rep in reps_summary.get("visitadores", []):
        lines.append(f"  {rep['nombre']}: ${rep['ventas_monto']:,.2f} — {rep['tasa_cumplimiento']}% cumplimiento")
    lines.append("")

    lines.append("TOP 10 MÉDICOS POR VENTAS")
    for doc in ranking.get("ranking", []):
        lines.append(f"  {doc['posicion']}. {doc['nombre']} ({doc['especialidad']}): ${doc['monto_mes']:,.2f}")
    lines.append("")

    report_text = "\n".join(lines)

    # Send via SMTP
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    report_email = os.getenv("REPORT_EMAIL")

    sent = False
    if smtp_user and smtp_password and report_email:
        try:
            msg = MIMEMultipart()
            msg["From"] = smtp_user
            msg["To"] = report_email
            msg["Subject"] = f"Reporte Semanal Narma — {now.strftime('%d/%m/%Y')}"
            msg.attach(MIMEText(report_text, "plain", "utf-8"))

            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.ehlo()
                server.starttls()
                server.login(smtp_user, smtp_password)
                server.sendmail(smtp_user, report_email, msg.as_string())
            sent = True
        except Exception as e:
            pass

    return {"sent": sent, "report": report_text}


# ── Chat endpoint ─────────────────────────────────────────────────────────────

@router.post("/chat", response_model=MikeChatResponse)
def mike_chat(request: MikeChatRequest, db: Session = Depends(get_db)):
    api_key = (os.getenv("ANTHROPIC_API_KEY") or "").strip()
    if not api_key:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY no configurada")

    try:
        return _run_mike_chat(request, db, api_key)
    except HTTPException:
        raise
    except anthropic.AuthenticationError:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY inválida o expirada")
    except anthropic.RateLimitError:
        raise HTTPException(status_code=429, detail="Límite de API alcanzado. Intenta en unos segundos.")
    except anthropic.APIConnectionError as e:
        raise HTTPException(status_code=503, detail=f"No se pudo conectar con la API de Anthropic: {str(e)}")
    except anthropic.BadRequestError as e:
        raise HTTPException(status_code=500, detail=f"Error en la solicitud a la API: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno: {type(e).__name__}: {str(e)}")


def _run_mike_chat(request: MikeChatRequest, db: Session, api_key: str) -> MikeChatResponse:
    client = anthropic.Anthropic(api_key=api_key)

    messages = []
    for msg in request.conversation_history:
        messages.append({"role": msg.role, "content": msg.content})
    messages.append({"role": "user", "content": request.message})

    # Cargar memorias persistentes e inyectarlas en el system prompt
    memories = db.query(MikeMemory).order_by(MikeMemory.created_at.desc()).limit(30).all()
    memory_block = ""
    if memories:
        lines = [f"[{m.id}|{m.category}] {m.content}" for m in reversed(memories)]
        memory_block = "\n\n## Memoria persistente (lo que recuerdas de sesiones anteriores)\n" + "\n".join(lines)

    system = MIKE_SYSTEM_PROMPT.format(current_date=datetime.utcnow().strftime("%d/%m/%Y %H:%M")) + memory_block

    final_response = ""
    max_iterations = 8
    iteration = 0

    # Track tool calls and results for chart building
    tool_calls_results: list = []
    export_url: Optional[str] = None

    while iteration < max_iterations:
        iteration += 1
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4096,
            system=system,
            tools=MIKE_TOOLS,
            messages=messages
        )

        if response.stop_reason == "end_turn":
            for block in response.content:
                if hasattr(block, "text"):
                    final_response = block.text
            break

        elif response.stop_reason == "tool_use":
            tool_results = []
            assistant_content = response.content

            for block in response.content:
                if block.type == "tool_use":
                    tool_result = execute_mike_tool(block.name, block.input, db)

                    # Track tool calls for chart building
                    tool_calls_results.append({"name": block.name, "input": block.input, "result": tool_result})

                    # Check for export — use path without /api prefix so frontend
                    # can call it directly via the axios instance (baseURL already has /api)
                    if block.name == "export_to_excel" and tool_result.get("download_ready"):
                        token = tool_result.get("token")
                        export_url = f"/mike/export/{token}"

                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": json.dumps(tool_result, ensure_ascii=False, default=str)
                    })

            messages.append({
                "role": "assistant",
                "content": [
                    {"type": b.type, **({
                        "text": b.text
                    } if b.type == "text" else {
                        "id": b.id,
                        "name": b.name,
                        "input": b.input
                    })}
                    for b in assistant_content
                ]
            })
            messages.append({"role": "user", "content": tool_results})
        else:
            final_response = "Ocurrió un error inesperado."
            break

    if not final_response:
        final_response = "No pude procesar tu solicitud."

    # Build charts from tool results
    charts = []
    for tc in tool_calls_results:
        name = tc["name"]
        result = tc["result"]

        if name == "get_monthly_trend" and "tendencia" in result:
            charts.append({
                "type": "line",
                "title": "Tendencia mensual",
                "data": result["tendencia"],
                "xKey": "mes",
                "yKey": "ventas_monto",
                "yKey2": "visitas_completadas"
            })

        elif name == "get_doctor_ranking" and "ranking" in result:
            charts.append({
                "type": "bar",
                "title": "Ranking médicos",
                "data": result["ranking"][:15],
                "xKey": "nombre",
                "yKey": "unidades"
            })

        elif name == "get_all_reps_summary" and "visitadores" in result:
            charts.append({
                "type": "bar",
                "title": "Desempeño visitadores",
                "data": result["visitadores"],
                "xKey": "nombre",
                "yKey": "tasa_cumplimiento"
            })

        elif name == "get_visits_tracking" and "tracking" in result:
            charts.append({
                "type": "bar",
                "title": "Cumplimiento visitas",
                "data": result["tracking"],
                "xKey": "rep_name",
                "yKey": "completion_rate"
            })

        elif name == "get_rep_commissions" and "comisiones" in result:
            charts.append({
                "type": "bar",
                "title": "Comisiones por visitador",
                "data": result["comisiones"],
                "xKey": "nombre",
                "yKey": "total_ventas"
            })

        elif name == "get_patient_analysis":
            analysis_type = tc["input"].get("analysis_type", "ranking")
            if analysis_type == "ranking" and "ranking" in result:
                charts.append({
                    "type": "bar",
                    "title": "Ranking de pacientes",
                    "data": result["ranking"][:15],
                    "xKey": "nombre",
                    "yKey": "total_monto"
                })
            elif analysis_type == "by_category" and "por_categoria" in result:
                charts.append({
                    "type": "bar",
                    "title": "Pacientes por categoría de producto",
                    "data": result["por_categoria"],
                    "xKey": "categoria",
                    "yKey": "pacientes_unicos"
                })

        elif name == "get_patient_detail" and "evolucion_mensual" in result:
            charts.append({
                "type": "line",
                "title": f"Evolución de compras — {result.get('nombre', 'Paciente')}",
                "data": result["evolucion_mensual"],
                "xKey": "mes",
                "yKey": "monto"
            })

    updated_history = list(request.conversation_history)
    updated_history.append(AgentMessage(role="user", content=request.message))
    updated_history.append(AgentMessage(role="assistant", content=final_response))

    return MikeChatResponse(
        response=final_response,
        conversation_history=updated_history,
        charts=charts,
        export_url=export_url
    )


# ── Memory endpoints ──────────────────────────────────────────────────────────

@router.get("/memory")
def get_mike_memory(db: Session = Depends(get_db)):
    mems = db.query(MikeMemory).order_by(MikeMemory.created_at.desc()).all()
    return [
        {"id": m.id, "content": m.content, "category": m.category,
         "created_at": m.created_at.isoformat() if m.created_at else None}
        for m in mems
    ]


@router.delete("/memory/{memory_id}")
def delete_mike_memory(memory_id: int, db: Session = Depends(get_db)):
    mem = db.query(MikeMemory).filter(MikeMemory.id == memory_id).first()
    if not mem:
        raise HTTPException(status_code=404, detail="Recuerdo no encontrado")
    db.delete(mem)
    db.commit()
    return {"ok": True}
